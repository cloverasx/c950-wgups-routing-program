# TODO: Update student ID# and remove this comment to ensure the line below
#       is the first line in this file.
# Name: A. Wayne Roberts, Student ID: C
# Course: C950 - Data Structures and Algorithms

# TODO: explain the process and the flow of the program

from datetime import datetime

# local imports
from algorithms.routing_algorithm import Truck
from algorithms.routing_algorithm import RoutingAlgorithm
from data_structures.package import Package
from data_structures.hash_table import HashTable
from ui.command_line_interface import CommandLineInterface
from utils.distance_calculator import DistanceCalculator
from utils.time_utils import TimeUtils


# TODO: Test function after parsing function works
def load_package_data(file_path):
    # open the file in read mode
    with open(file_path, "r") as file:
        # read the lines from the file
        lines = file.readlines()
        # create an empty list to store the packages
        packages = []
        # iterate through the lines
        for line in lines:
            # split the line by the comma
            data = line.strip().split(",")
            # extract the package data
            package_id = int(data[0])
            address = data[1]
            city = data[2]
            state = data[3]
            zip_code = data[4]
            deadline = data[5]
            weight = data[6]
            status = "At Hub"
            # create a package object
            package = Package(
                package_id, address, city, state, zip_code, deadline, weight, status
            )
            # add the package to the list
            packages.append(package)
        # return the list of packages
        return packages


# TODO: Test function after parsing function works
def load_distance_table(file_path):
    # open the file in read mode
    with open(file_path, "r") as file:
        # read the lines from the file
        lines = file.readlines()
        # create an empty dictionary to store the distances
        distances = {}
        # iterate through the lines
        for line in lines:
            # split the line by the comma
            data = line.strip().split(",")
            # extract the addresses and distances
            address1 = data[0]
            address2 = data[1]
            distance = float(data[2])
            # check if the address is already in the dictionary
            if address1 in distances:
                # add the distance to the existing address
                distances[address1][address2] = distance
            else:
                # create a new entry for the address
                distances[address1] = {address2: distance}
        # return the dictionary of distances
        return distances


# TODO: These libraries are associated with the next functions and are only temporary
# until they are placed in their proper locations in the project.
import pandas as pd
import csv
import re
import openpyxl
import re


# TODO: This should be moved to an appropriate location; it is only for use with the
# process_xlsx() function.
def clean_string(s):
    # Note: I was having problems finding the data in the xlsx file and found
    # that this could clean up the data.
    return re.sub(r"\s+", " ", str(s).strip().lower())


# TODO: the following works well with the package file currently. Will need updated for
# the distance table file and any logic checking and error handling.
def process_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    header_row = None
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=worksheet.max_row), start=1
    ):
        cell_values = [clean_string(cell.value) for cell in row]
        # print(f"Row {row_index}: {cell_values}")  # Debugging output
        if "address" in cell_values and "city" in cell_values:
            header_row = row_index
            # print(f"Header row found at row {header_row}")  # Debugging output
            break
    if header_row is None:
        raise ValueError(
            "Could not find a header row containing both 'address' and 'city'."
        )
    headers = [clean_string(cell.value) for cell in worksheet[header_row] if cell.value]
    # print(f"Headers: {headers}")  # Debugging output
    data = []
    for row in worksheet.iter_rows(
        min_row=header_row + 1, max_row=worksheet.max_row, values_only=True
    ):
        if any(cell for cell in row):
            data.append(
                [clean_string(cell) if cell else "" for cell in row[: len(headers)]]
            )
    return headers, data


# Debugging output
try:
    file_path = "data/WGUPS Package File.xlsx"
    package_headers, package_data = process_xlsx(file_path)
    print("Number of data rows:", len(package_data))
    print("First data row:", package_data[0])
except Exception as e:
    print(f"An error occurred: {str(e)}")

# Debugging output
with open("temp.csv", "w", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(package_headers)
    for row in package_data:
        writer.writerow(row)


import openpyxl
import re


# TODO: this is going to need to be cleaned up for the output, but the current version 
# works for now.
def process_distance_table(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    # Find the last row with data in all columns
    last_full_row = None
    for row in reversed(list(worksheet.rows)):
        if all(cell.value is not None for cell in row):
            last_full_row = row
            break

    if last_full_row is None:
        raise ValueError("Could not find a fully populated row in the distance table.")

    # Determine the number of columns in the matrix
    matrix_size = sum(1 for cell in last_full_row if cell.value is not None)

    # Calculate the start row
    start_row = last_full_row[0].row - matrix_size + 1

    # Extract location names and distances
    locations = []
    distances = []
    for row in worksheet.iter_rows(
        min_row=start_row, max_row=last_full_row[0].row, values_only=True
    ):
        location_name = clean_string(row[0])
        locations.append(location_name)
        # Start from index 1 to skip the first column (location name)
        row_distances = [
            float(cell) if isinstance(cell, (int, float)) else 0
            for cell in row[1:matrix_size]
        ]
        distances.append(row_distances)

    return locations, distances


# Usage
try:
    file_path = "data/WGUPS Distance Table.xlsx"
    locations, distances = process_distance_table(file_path)
    print(f"Number of locations: {len(locations)}")
    print(f"First location: {locations[0]}")
    print(f"Last location: {locations[-1]}")
    print(f"Size of distance matrix: {len(distances)}x{len(distances[0])}")
    print(f"First row of distances: {distances[0]}")
except Exception as e:
    print(f"An error occurred: {str(e)}")
