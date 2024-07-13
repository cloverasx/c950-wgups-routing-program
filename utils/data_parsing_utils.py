import openpyxl
import csv
from utils.string_utils import StringUtils as su


class DataParsingUtils:
    @staticmethod
    def clean_string(s):
        # Note: I was having problems finding the data in the xlsx file and found
        # that this could clean up the data.
        s = str(s)
        s = su.remove_commas(s)
        s = su.clean_whitespace(s)
        s = su.convert_to_lowercase(s)
        return s

    # TODO: the following works well with the package file currently. Will need updated for
    # the distance table file and any logic checking and error handling.
    @staticmethod
    def process_package_file(file_path):
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        header_row = None
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=worksheet.max_row), start=1
        ):
            cell_values = [DataParsingUtils.clean_string(cell.value) for cell in row]
            # print(f"Row {row_index}: {cell_values}")  # Debugging output
            if "address" in cell_values and "city" in cell_values:
                header_row = row_index
                # print(f"Header row found at row {header_row}")  # Debugging output
                break
        if header_row is None:
            raise ValueError(
                "Could not find a header row containing both 'address' and 'city'."
            )
        headers = [
            DataParsingUtils.clean_string(cell.value)
            for cell in worksheet[header_row]
            if cell.value
        ]
        # print(f"Headers: {headers}")  # Debugging output
        data = []
        for row in worksheet.iter_rows(
            min_row=header_row + 1, max_row=worksheet.max_row, values_only=True
        ):
            if any(cell for cell in row):
                data.append(
                    [
                        DataParsingUtils.clean_string(cell) if cell else ""
                        for cell in row[: len(headers)]
                    ]
                )
        return headers, data

    @staticmethod
    def save_package_file(package_headers, package_data):
        with open("data/package_file.csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(package_headers)
            for row in package_data:
                writer.writerow(row)

    @staticmethod
    def process_distance_table(file_path):
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active

        # Find the last row with data in all columns
        last_full_row = None
        for row in reversed(list(worksheet.rows)):
            if all(cell.value is not None for cell in row):
                last_full_row = row
                break

        # debug: print the row number for the last full row
        print(f"Last full row: {last_full_row[0].row}")

        if last_full_row is None:
            raise ValueError(
                "Could not find a fully populated row in the distance table."
            )

        # Determine the valid data range
        matrix_size = len(last_full_row)
        start_row = last_full_row[0].row - (matrix_size - 3)  # 0-index + headers

        # Extract location names and distances
        locations = []
        distances = []
        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=start_row, max_row=last_full_row[0].row, values_only=True
            ),
            start=0,
        ):
            location_name = next(
                (
                    line.strip()
                    for line in str(row[0]).split("\n")
                    if line.strip() and line.strip()[0].isdigit()
                ),
                "",
            )
            street_address = DataParsingUtils.clean_string(location_name)
            locations.append(street_address)

            row_distances = []
            for col_index in range(2, matrix_size):
                cell = row[col_index]
                if isinstance(cell, (int, float)):
                    row_distances.append(float(cell))
                else:
                    # If cell is empty, get the symmetric value
                    symmetric_cell = worksheet.cell(
                        row=start_row + col_index - 2, column=row_index + 3
                    ).value
                    row_distances.append(float(symmetric_cell) if symmetric_cell else 0)

            distances.append(row_distances)

        return locations, distances

    @staticmethod
    def save_distance_file(locations, distances, output_file):
        with open(output_file, "w", newline="") as file:
            writer = csv.writer(file)
            # Write the header row
            writer.writerow(
                # ["location"] + [f"Distance_{i+1}" for i in range(len(locations))]
                ["location"] + locations
            )
            for i, distances_row in enumerate(distances):
                writer.writerow([locations[i]] + distances_row)
